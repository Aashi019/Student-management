
from flask import Flask, render_template, redirect, url_for, request, flash, jsonify, send_file, Response
from flask_socketio import SocketIO, emit
from flask_login import LoginManager, login_user, login_required, logout_user, current_user
from models import db, User, Student, Subject, Grade, Attendance, Enrollment, AcademicYear, Teacher, Assignment, AssignmentTemplate, SubjectTeacher, FeeStructure, FeePayment, FeeReceipt
from forms import LoginForm, StudentForm, GradeForm, AttendanceForm, SubjectForm
from werkzeug.utils import secure_filename
import os
import json
import time
from datetime import datetime, timedelta
from sqlalchemy import func, desc, case
import io
import csv
from io import StringIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Initialize Flask app
app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-change-in-production'

# Database configuration - Force PostgreSQL usage (no SQLite fallback)
DATABASE_URL = 'postgresql://neondb_owner:npg_XCrcYpHg9SO8@ep-blue-snow-ah4jpxiz-pooler.c-3.us-east-1.aws.neon.tech/neondb?sslmode=require'

# Always use Neon PostgreSQL
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL', DATABASE_URL)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['WTF_CSRF_ENABLED'] = True
app.config['WTF_CSRF_TIME_LIMIT'] = None

# Initialize extensions
db.init_app(app)
# SocketIO for Vercel serverless compatibility
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='eventlet', logger=False, engineio_logger=False)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'

# File upload configuration
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Ensure upload directory exists (skip in serverless environment)
try:
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
except:
    pass  # Skip in serverless environment

# Register API Blueprint (with error handling for serverless)
try:
    from api_routes import api_bp, set_socketio
    app.register_blueprint(api_bp)
    set_socketio(socketio)  # Pass socketio instance to api_routes
except ImportError as e:
    print(f"Warning: Could not import api_routes: {e}")
except Exception as e:
    print(f"Warning: Error setting up API routes: {e}")

@login_manager.user_loader
def load_user(user_id):
    # Support both User and Student login
    if isinstance(user_id, str) and user_id.startswith('student_'):
        student_id = int(user_id.replace('student_', ''))
        student = Student.query.get(student_id)
        if student:
            # Make student compatible with Flask-Login
            student.is_authenticated = True
            student.is_active = (student.status == 'active')
            student.is_anonymous = False
        return student
    return User.query.get(int(user_id))

# Role-based authorization decorators
from functools import wraps
from flask import abort

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            flash('Administrator access required.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def teacher_or_admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role not in ['admin', 'teacher']:
            flash('Teacher or Administrator access required.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def student_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'student':
            flash('Student access required.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def get_teacher_students(teacher_user_id):
    """Get all students that a teacher teaches through subject assignments"""
    teacher = Teacher.query.filter_by(user_id=teacher_user_id).first()
    if not teacher:
        return []
    
    # Get subjects that this teacher teaches
    teacher_subjects = db.session.query(SubjectTeacher.subject_id).filter_by(teacher_id=teacher.id).subquery()
    
    # Get students enrolled in those subjects
    student_ids = db.session.query(Enrollment.student_id).filter(
        Enrollment.subject_id.in_(teacher_subjects),
        Enrollment.status == 'enrolled'
    ).distinct().subquery()
    
    return Student.query.filter(Student.id.in_(student_ids)).all()

def can_teacher_access_student(teacher_user_id, student_id):
    """Check if a teacher can access a specific student's data"""
    teacher = Teacher.query.filter_by(user_id=teacher_user_id).first()
    if not teacher:
        return False
    
    # Get subjects that this teacher teaches
    teacher_subjects = db.session.query(SubjectTeacher.subject_id).filter_by(teacher_id=teacher.id).subquery()
    
    # Check if student is enrolled in any of those subjects
    enrollment = Enrollment.query.filter(
        Enrollment.student_id == student_id,
        Enrollment.subject_id.in_(teacher_subjects),
        Enrollment.status == 'enrolled'
    ).first()
    
    return enrollment is not None

# Authentication Routes
@app.route('/')
def home():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        # Redirect based on user type
        if hasattr(current_user, 'student_id'):
            return redirect(url_for('student_dashboard'))
        return redirect(url_for('dashboard'))
    
    form = LoginForm()
    if form.validate_on_submit():
        # Try to login as User (admin/teacher/staff)
        user = User.query.filter_by(username=form.username.data).first()
        if user and user.check_password(form.password.data):
            login_user(user, remember=form.remember_me.data)
            user.last_login = datetime.utcnow()
            db.session.commit()
            flash('Logged in successfully!', 'success')
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('dashboard'))
        
        # Try to login as Student (using student_id)
        student = Student.query.filter_by(student_id=form.username.data).first()
        if student and student.check_password(form.password.data):
            student.is_authenticated = True
            student.is_active = (student.status == 'active')
            student.is_anonymous = False
            login_user(student, remember=form.remember_me.data)
            flash('Welcome to Student Portal!', 'success')
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('student_dashboard'))
        
        flash('Invalid username or password.', 'danger')
    return render_template('login.html', form=form)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))

# Main Dashboard
@app.route('/dashboard')
@login_required
def dashboard():
    # Check if user is a student
    if current_user.role == 'student':
        # Render simple student dashboard
        return render_template('dashboard.html')
    
    # For admin/teacher users, show full dashboard with statistics
    # Get statistics for dashboard
    total_students = Student.query.filter_by(status='active').count()
    total_subjects = Subject.query.count()
    
    # Get recent enrollments (last 30 days)
    thirty_days_ago = datetime.now() - timedelta(days=30)
    recent_enrollments = Student.query.filter(Student.enrollment_date >= thirty_days_ago).count()
    
    # Get attendance rate for current month
    current_month = datetime.now().month
    current_year = datetime.now().year
    
    # Simple approach: count total and present separately
    total_attendance = Attendance.query.filter(
        func.extract('month', Attendance.date) == current_month,
        func.extract('year', Attendance.date) == current_year
    ).count()
    
    present_attendance = Attendance.query.filter(
        func.extract('month', Attendance.date) == current_month,
        func.extract('year', Attendance.date) == current_year,
        Attendance.status == 'present'
    ).count()
    
    attendance_rate = 0
    if total_attendance and total_attendance > 0:
        attendance_rate = round((present_attendance / total_attendance) * 100, 1)
    
    # Get grade distribution
    grade_distribution = db.session.query(
        Grade.letter_grade,
        func.count(Grade.id).label('count')
    ).group_by(Grade.letter_grade).all()

    return render_template('dashboard/index.html', 
                         total_students=total_students,
                         total_subjects=total_subjects,
                         recent_enrollments=recent_enrollments,
                         attendance_rate=attendance_rate,
                         grade_distribution=grade_distribution)

# Student Management Routes
@app.route('/students')
@login_required
@teacher_or_admin_required
def students():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    grade_filter = request.args.get('grade', '', type=str)
    status_filter = request.args.get('status', 'active', type=str)
    
    # Teachers can view all students, admins can see all students  
    # (editing restrictions are handled at the route level)
    query = Student.query
    
    if search:
        query = query.filter(
            (Student.first_name.contains(search)) |
            (Student.last_name.contains(search)) |
            (Student.email.contains(search)) |
            (Student.student_id.contains(search))
        )
    
    if grade_filter:
        query = query.filter(Student.grade_level == grade_filter)
    
    if status_filter:
        query = query.filter(Student.status == status_filter)
    
    students_pagination = query.order_by(Student.last_name, Student.first_name).paginate(
        page=page, per_page=20, error_out=False
    )
    
    # Get unique grade levels for filter
    grade_levels = db.session.query(Student.grade_level).distinct().all()
    grade_levels = [g[0] for g in grade_levels if g[0]]
    
    return render_template('students/list.html', 
                         students=students_pagination.items,
                         pagination=students_pagination,
                         search=search,
                         grade_filter=grade_filter,
                         status_filter=status_filter,
                         grade_levels=grade_levels)

@app.route('/students/add', methods=['GET', 'POST'])
@login_required
@teacher_or_admin_required  # Allow teachers and admins to add students
def add_student():
    form = StudentForm()
    if form.validate_on_submit():
        # Generate unique student ID
        last_student = Student.query.order_by(desc(Student.id)).first()
        next_id = (last_student.id + 1) if last_student else 1
        student_id = f"STU{next_id:06d}"
        
        student = Student(
            student_id=student_id,
            first_name=form.first_name.data,
            last_name=form.last_name.data,
            email=form.email.data,
            phone=form.phone.data,
            date_of_birth=form.date_of_birth.data,
            gender=form.gender.data,
            address=form.address.data,
            city=form.city.data,
            state=form.state.data,
            zip_code=form.zip_code.data,
            emergency_contact=form.emergency_contact.data,
            emergency_phone=form.emergency_phone.data,
            grade_level=form.grade_level.data,
            graduation_year=form.graduation_year.data
        )
        
        # Handle photo upload
        if form.photo.data:
            filename = secure_filename(form.photo.data.filename)
            if filename:
                photo_filename = f"{student_id}_{filename}"
                form.photo.data.save(os.path.join(app.config['UPLOAD_FOLDER'], photo_filename))
                student.photo_filename = photo_filename
        
        db.session.add(student)
        db.session.commit()
        
        # Emit real-time update
        notify_student_created(student)
        
        flash('Student added successfully!', 'success')
        return redirect(url_for('students'))
    
    return render_template('students/form.html', form=form, title='Add Student')

@app.route('/students/<int:id>')
@login_required
@teacher_or_admin_required
def student_detail(id):
    student = Student.query.get_or_404(id)
    
    # Check if teacher can access this student
    if current_user.role == 'teacher' and not can_teacher_access_student(current_user.id, id):
        flash('You can only view students you teach.', 'error')
        return redirect(url_for('students'))
    
    # Get student's grades
    grades = Grade.query.filter_by(student_id=id).join(Subject).all()
    
    # Get attendance records (last 30 days)
    thirty_days_ago = datetime.now().date() - timedelta(days=30)
    attendance = Attendance.query.filter(
        Attendance.student_id == id,
        Attendance.date >= thirty_days_ago
    ).order_by(desc(Attendance.date)).all()
    
    # Calculate attendance rate
    total_days = len(attendance)
    present_days = len([a for a in attendance if a.status == 'present'])
    attendance_rate = round((present_days / total_days * 100), 1) if total_days > 0 else 0
    
    return render_template('students/detail.html', 
                         student=student,
                         grades=grades,
                         attendance=attendance,
                         attendance_rate=attendance_rate)

@app.route('/students/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required  # Only admins can edit student personal data
def edit_student(id):
    student = Student.query.get_or_404(id)
    
    # Check if teacher can access this student
    if current_user.role == 'teacher' and not can_teacher_access_student(current_user.id, id):
        flash('You can only edit students you teach.', 'error')
        return redirect(url_for('students'))
    
    form = StudentForm(obj=student)
    
    if form.validate_on_submit():
        form.populate_obj(student)
        
        # Handle photo upload
        if form.photo.data:
            filename = secure_filename(form.photo.data.filename)
            if filename:
                photo_filename = f"{student.student_id}_{filename}"
                form.photo.data.save(os.path.join(app.config['UPLOAD_FOLDER'], photo_filename))
                student.photo_filename = photo_filename
        
        student.updated_at = datetime.utcnow()
        db.session.commit()
        
        # Emit real-time update
        notify_student_updated(student)
        
        flash('Student updated successfully!', 'success')
        return redirect(url_for('student_detail', id=id))
    
    return render_template('students/form.html', form=form, student=student, title='Edit Student')

@app.route('/students/<int:id>/delete', methods=['POST'])
@login_required
@admin_required  # Only admins can delete students
def delete_student(id):
    student = Student.query.get_or_404(id)
    student.status = 'inactive'
    student.updated_at = datetime.utcnow()
    db.session.commit()
    
    # Emit real-time update
    notify_student_deleted(id)
    
    flash('Student deactivated successfully!', 'success')
    return redirect(url_for('students'))

# API Routes for Charts and Real-time Data
@app.route('/api/dashboard/stats')
@login_required
@teacher_or_admin_required
def api_dashboard_stats():
    # Student enrollment trend (last 12 months)
    enrollment_data = []
    for i in range(12):
        date = datetime.now() - timedelta(days=30*i)
        count = Student.query.filter(
            Student.enrollment_date <= date.date()
        ).count()
        enrollment_data.append({
            'month': date.strftime('%b %Y'),
            'count': count
        })
    
    # Grade distribution
    grade_dist = db.session.query(
        Grade.letter_grade,
        func.count(Grade.id).label('count')
    ).group_by(Grade.letter_grade).all()
    
    grade_data = [{'grade': g[0], 'count': g[1]} for g in grade_dist if g[0]]
    
    # Attendance trend (last 30 days)
    attendance_trend = []
    for i in range(30):
        date = datetime.now().date() - timedelta(days=i)
        total = Attendance.query.filter(Attendance.date == date).count()
        present = Attendance.query.filter(
            Attendance.date == date,
            Attendance.status == 'present'
        ).count()
        
        rate = (present / total * 100) if total > 0 else 0
        attendance_trend.append({
            'date': date.strftime('%Y-%m-%d'),
            'rate': round(rate, 1)
        })
    
    return jsonify({
        'enrollment_trend': list(reversed(enrollment_data)),
        'grade_distribution': grade_data,
        'attendance_trend': list(reversed(attendance_trend))
    })

@app.route('/api/students/chart-data')
@login_required
@teacher_or_admin_required
def api_students_chart_data():
    # Students by grade level
    grade_data = db.session.query(
        Student.grade_level,
        func.count(Student.id).label('count')
    ).filter(Student.status == 'active').group_by(Student.grade_level).all()
    
    # Students by gender
    gender_data = db.session.query(
        Student.gender,
        func.count(Student.id).label('count')
    ).filter(Student.status == 'active').group_by(Student.gender).all()
    
    return jsonify({
        'by_grade': [{'grade': g[0], 'count': g[1]} for g in grade_data],
        'by_gender': [{'gender': g[0], 'count': g[1]} for g in gender_data]
    })

# Export Routes
@app.route('/export/students')
@login_required
@teacher_or_admin_required
def export_students():
    format_type = request.args.get('format', 'excel')
    search = request.args.get('search', '')
    grade_filter = request.args.get('grade', '')
    status_filter = request.args.get('status', 'active')
    student_ids = request.args.getlist('student_ids')
    
    # Build query
    query = Student.query
    
    # Apply filters
    if search:
        query = query.filter(
            db.or_(
                Student.first_name.contains(search),
                Student.last_name.contains(search),
                Student.email.contains(search),
                Student.student_id.contains(search)
            )
        )
    
    if grade_filter:
        query = query.filter(Student.grade_level == int(grade_filter))
    
    if status_filter:
        query = query.filter(Student.status == status_filter)
    
    # If specific student IDs are provided, filter by them
    if student_ids:
        query = query.filter(Student.id.in_(student_ids))
    
    students = query.all()
    
    if not students:
        flash('No students found for export.', 'warning')
        return redirect(url_for('students'))
    
    if format_type == 'csv':
        # CSV export
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            'Student ID', 'First Name', 'Last Name', 'Full Name', 'Email', 
            'Phone', 'Grade Level', 'GPA', 'Status', 'Gender', 'Date of Birth',
            'Address', 'City', 'State', 'Zip Code', 'Enrollment Date', 
            'Emergency Contact', 'Emergency Phone', 'Parent Email'
        ])
        
        # Write data
        for student in students:
            writer.writerow([
                student.student_id,
                student.first_name,
                student.last_name,
                student.full_name,
                student.email,
                student.phone or '',
                student.grade_level,
                f"{student.gpa:.2f}" if student.gpa else '',
                student.status,
                student.gender or '',
                student.date_of_birth.strftime('%Y-%m-%d') if student.date_of_birth else '',
                student.address or '',
                student.city or '',
                student.state or '',
                student.zip_code or '',
                student.enrollment_date.strftime('%Y-%m-%d') if student.enrollment_date else '',
                student.emergency_contact or '',
                student.emergency_phone or '',
                student.parent_email or ''
            ])
        
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename=students_{datetime.now().strftime("%Y%m%d")}.csv'}
        )
        
    elif format_type == 'excel':
        # Excel export using openpyxl directly
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Students"
            
            # Headers
            headers = [
                'Student ID', 'First Name', 'Last Name', 'Full Name', 'Email', 
                'Phone', 'Grade Level', 'GPA', 'Status', 'Gender', 'Date of Birth',
                'Address', 'City', 'State', 'Zip Code', 'Enrollment Date', 
                'Emergency Contact', 'Emergency Phone', 'Parent Email'
            ]
            
            # Style for headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Add headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Add data
            for row_num, student in enumerate(students, 2):
                data = [
                    student.student_id,
                    student.first_name,
                    student.last_name,
                    student.full_name,
                    student.email,
                    student.phone or '',
                    student.grade_level,
                    f"{student.gpa:.2f}" if student.gpa else '',
                    student.status.title(),
                    student.gender or '',
                    student.date_of_birth.strftime('%Y-%m-%d') if student.date_of_birth else '',
                    student.address or '',
                    student.city or '',
                    student.state or '',
                    student.zip_code or '',
                    student.enrollment_date.strftime('%Y-%m-%d') if student.enrollment_date else '',
                    student.emergency_contact or '',
                    student.emergency_phone or '',
                    student.parent_email or ''
                ]
                
                for col_num, value in enumerate(data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Save to BytesIO
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return send_file(
                output,
                as_attachment=True,
                download_name=f'students_{datetime.now().strftime("%Y%m%d")}.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        except Exception as e:
            flash(f'Error generating Excel file: {str(e)}', 'error')
            return redirect(url_for('students'))
    
    elif format_type == 'pdf':
        # PDF export
        try:
            output = io.BytesIO()
            doc = SimpleDocTemplate(output, pagesize=A4)
            story = []
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = styles['Title']
            heading_style = styles['Heading2']
            normal_style = styles['Normal']
            
            # Title
            title = Paragraph("Student List Report", title_style)
            story.append(title)
            story.append(Spacer(1, 20))
            
            # Report info
            report_info = Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}<br/>Total Students: {len(students)}", normal_style)
            story.append(report_info)
            story.append(Spacer(1, 20))
            
            # Table data
            data = [['Student ID', 'Name', 'Email', 'Grade', 'GPA', 'Status']]
            for student in students:
                data.append([
                    student.student_id,
                    student.full_name,
                    student.email,
                    f"Grade {student.grade_level}",
                    f"{student.gpa:.2f}" if student.gpa else 'N/A',
                    student.status.title()
                ])
            
            # Create table
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
            doc.build(story)
            output.seek(0)
            
            return send_file(
                output,
                as_attachment=True,
                download_name=f'students_{datetime.now().strftime("%Y%m%d")}.pdf',
                mimetype='application/pdf'
            )
        except Exception as e:
            flash(f'Error generating PDF file: {str(e)}', 'error')
            return redirect(url_for('students'))
    
    elif format_type == 'json':
        # JSON export
        students_data = []
        for student in students:
            students_data.append({
                'student_id': student.student_id,
                'first_name': student.first_name,
                'last_name': student.last_name,
                'full_name': student.full_name,
                'email': student.email,
                'phone': student.phone,
                'grade_level': student.grade_level,
                'gpa': student.gpa,
                'status': student.status,
                'gender': student.gender,
                'date_of_birth': student.date_of_birth.strftime('%Y-%m-%d') if student.date_of_birth else None,
                'address': student.address,
                'city': student.city,
                'state': student.state,
                'zip_code': student.zip_code,
                'enrollment_date': student.enrollment_date.strftime('%Y-%m-%d') if student.enrollment_date else None,
                'emergency_contact': student.emergency_contact,
                'emergency_phone': student.emergency_phone,
                'parent_email': student.parent_email
            })
        
        return Response(
            json.dumps(students_data, indent=2),
            mimetype='application/json',
            headers={'Content-Disposition': f'attachment; filename=students_{datetime.now().strftime("%Y%m%d")}.json'}
        )
    
    else:
        flash('Invalid export format requested.', 'error')
        return redirect(url_for('students'))

# API Routes for Reports
@app.route('/api/students/export')
@login_required
@teacher_or_admin_required
def api_export_students():
    """API endpoint for student export - redirects to main export function"""
    return export_students()

@app.route('/api/subjects/export')
@login_required
@teacher_or_admin_required
def api_export_subjects():
    """Export subjects data"""
    format_type = request.args.get('format', 'csv')
    
    subjects = Subject.query.all()
    
    if format_type == 'csv':
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['Faculty ID', 'Faculty Name', 'Faculty Code', 'Description', 'Total Students', 'Department', 'Active'])
        
        # Write data
        for subject in subjects:
            # Count enrolled students
            enrolled_count = Enrollment.query.filter_by(subject_id=subject.id, status='enrolled').count()
            
            writer.writerow([
                subject.id,
                subject.name,
                subject.code or 'N/A',
                subject.description or 'N/A',
                enrolled_count,
                subject.department or 'N/A',
                'Yes' if subject.is_active else 'No'
            ])
        
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename=facultyreport_{datetime.now().strftime("%Y%m%d")}.csv'}
        )
    
    elif format_type == 'json':
        subjects_data = []
        for subject in subjects:
            # Count enrolled students
            enrolled_count = Enrollment.query.filter_by(subject_id=subject.id, status='enrolled').count()
            
            subjects_data.append({
                'id': subject.id,
                'faculty_name': subject.name,
                'faculty_code': subject.code,
                'description': subject.description,
                'total_students': enrolled_count,
                'department': subject.department,
                'is_active': subject.is_active
            })
        
        return Response(
            json.dumps(subjects_data, indent=2),
            mimetype='application/json',
            headers={'Content-Disposition': f'attachment; filename=facultyreport_{datetime.now().strftime("%Y%m%d")}.json'}
        )
    
    else:
        return jsonify({'error': 'Invalid format'}), 400

@app.route('/api/attendance/export')
@login_required
@teacher_or_admin_required
def api_export_attendance():
    """API endpoint for attendance export - redirects to main export function"""
    return export_attendance()

@app.route('/export/grades')
@app.route('/api/grades/export')  # Alternative route for compatibility
@login_required
@teacher_or_admin_required
def export_grades():
    """Export grades data"""
    format_type = request.args.get('format', 'csv')
    
    # Get all grades with student and subject info
    grades = Grade.query.join(Student).join(Subject).all()
    
    if format_type == 'csv':
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['Student ID', 'Student Name', 'Subject', 'Grade Value', 'Letter Grade', 'Grade Type', 'Semester', 'Date Recorded'])
        
        # Write data
        for grade in grades:
            writer.writerow([
                grade.student.student_id,
                grade.student.full_name,
                grade.subject.name,
                grade.grade_value,
                grade.letter_grade,
                grade.grade_type,
                grade.semester,
                grade.date_recorded.strftime('%Y-%m-%d') if grade.date_recorded else ''
            ])
        
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename=grades_{datetime.now().strftime("%Y%m%d")}.csv'}
        )
    
    elif format_type == 'json':
        grades_data = []
        for grade in grades:
            grades_data.append({
                'student_id': grade.student.student_id,
                'student_name': grade.student.full_name,
                'subject': grade.subject.name,
                'grade_value': grade.grade_value,
                'letter_grade': grade.letter_grade,
                'grade_type': grade.grade_type,
                'semester': grade.semester,
                'date_recorded': grade.date_recorded.strftime('%Y-%m-%d') if grade.date_recorded else None
            })
        
        return Response(
            json.dumps(grades_data, indent=2),
            mimetype='application/json',
            headers={'Content-Disposition': f'attachment; filename=grades_{datetime.now().strftime("%Y%m%d")}.json'}
        )
    
    else:
        return jsonify({'error': 'Invalid format'}), 400

@app.route('/export/attendance')
@app.route('/api/attendances/export')  # Alternative route for compatibility
@login_required
@teacher_or_admin_required
def export_attendance():
    """Export attendance data"""
    format_type = request.args.get('format', 'csv')
    
    # Get all attendance records with student info
    attendance_records = Attendance.query.join(Student).order_by(Attendance.date.desc()).all()
    
    if format_type == 'csv':
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['Student ID', 'Student Name', 'Date', 'Status', 'Period', 'Notes'])
        
        # Write data
        for record in attendance_records:
            writer.writerow([
                record.student.student_id,
                record.student.full_name,
                record.date.strftime('%Y-%m-%d'),
                record.status,
                record.period or '',
                record.notes or ''
            ])
        
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename=attendance_{datetime.now().strftime("%Y%m%d")}.csv'}
        )
    
    elif format_type == 'json':
        attendance_data = []
        for record in attendance_records:
            attendance_data.append({
                'student_id': record.student.student_id,
                'student_name': record.student.full_name,
                'date': record.date.strftime('%Y-%m-%d'),
                'status': record.status,
                'period': record.period,
                'notes': record.notes
            })
        
        return Response(
            json.dumps(attendance_data, indent=2),
            mimetype='application/json',
            headers={'Content-Disposition': f'attachment; filename=attendance_{datetime.now().strftime("%Y%m%d")}.json'}
        )
    
    else:
        return jsonify({'error': 'Invalid format'}), 400

# SocketIO Events for Real-time Updates
@socketio.on('connect')
def handle_connect():
    if current_user.is_authenticated:
        emit('status', {'msg': f'{current_user.username} has connected'}, broadcast=True)
        print(f'Client connected: {current_user.username}')

@socketio.on('disconnect')
def handle_disconnect():
    if current_user.is_authenticated:
        emit('status', {'msg': f'{current_user.username} has disconnected'}, broadcast=True)
        print(f'Client disconnected: {current_user.username}')

@socketio.on('request_dashboard_update')
def handle_dashboard_update():
    """Client requests dashboard data update"""
    if current_user.is_authenticated:
        # Get fresh stats
        total_students = Student.query.filter_by(status='active').count()
        total_subjects = Subject.query.filter_by(is_active=True).count()
        
        emit('dashboard_updated', {
            'total_students': total_students,
            'total_subjects': total_subjects,
            'timestamp': datetime.now().isoformat()
        })

# Helper function to broadcast updates
def broadcast_update(event_name, data):
    """Broadcast an update to all connected clients"""
    socketio.emit(event_name, data, namespace='/')

# Real-time notification functions (to be called after DB operations)
def notify_student_created(student):
    broadcast_update('student_created', {
        'id': student.id,
        'student_id': student.student_id,
        'name': student.full_name,
        'grade': student.grade_level
    })

def notify_student_updated(student):
    broadcast_update('student_updated', {
        'id': student.id,
        'student_id': student.student_id,
        'name': student.full_name,
        'grade': student.grade_level
    })

def notify_student_deleted(student_id):
    broadcast_update('student_deleted', {'student_id': student_id})

def notify_grade_added(grade):
    broadcast_update('grade_added', {
        'id': grade.id,
        'student_id': grade.student_id,
        'subject_id': grade.subject_id,
        'grade_value': grade.grade_value
    })

def notify_attendance_recorded(attendance):
    broadcast_update('attendance_recorded', {
        'id': attendance.id,
        'student_id': attendance.student_id,
        'date': attendance.date.isoformat(),
        'status': attendance.status
    })

def notify_subject_created(subject):
    broadcast_update('subject_created', {
        'id': subject.id,
        'name': subject.name,
        'code': subject.code
    })

def notify_subject_updated(subject):
    broadcast_update('subject_updated', {
        'id': subject.id,
        'name': subject.name,
        'code': subject.code
    })

# Subject Management Routes
@app.route('/subjects')
@login_required
@teacher_or_admin_required
def subjects():
    subjects = Subject.query.filter_by(is_active=True).all()
    
    # Calculate enrolled students for each faculty
    faculty_enrollments = []
    for subject in subjects:
        # Count students based on their grades (students who have grades in this faculty)
        enrolled_count = db.session.query(Student.id).join(Grade).filter(
            Grade.subject_id == subject.id,
            Student.status == 'active'
        ).distinct().count()
        
        # Also count students by faculty assignment in sample data
        # For BSC-CSIT faculty
        if subject.code == 'BSC-CSIT':
            enrolled_count = Student.query.filter(
                Student.status == 'active',
                Student.student_id.in_(['STU000001', 'STU000003', 'STU000005', 'STU000007', 'STU000009'])
            ).count()
        # For BCA faculty
        elif subject.code == 'BCA':
            enrolled_count = Student.query.filter(
                Student.status == 'active',
                Student.student_id.in_(['STU000002', 'STU000004', 'STU000006', 'STU000008', 'STU000010'])
            ).count()
        
        faculty_enrollments.append({
            'subject': subject,
            'enrolled_students': enrolled_count
        })
    
    return render_template('subjects/list.html', faculty_enrollments=faculty_enrollments)

@app.route('/subjects/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_subject():
    form = SubjectForm()
    if form.validate_on_submit():
        subject = Subject(
            name=form.name.data,
            code=form.code.data,
            description=form.description.data,
            credits=form.credits.data,
            department=form.department.data
        )
        db.session.add(subject)
        db.session.commit()
        
        notify_subject_created(subject)
        flash('Subject added successfully!', 'success')
        return redirect(url_for('subjects'))
    
    return render_template('subjects/form.html', form=form, title='Add Subject')

@app.route('/subjects/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_subject(id):
    subject = Subject.query.get_or_404(id)
    form = SubjectForm(obj=subject)
    
    if form.validate_on_submit():
        form.populate_obj(subject)
        subject.updated_at = datetime.utcnow()
        db.session.commit()
        
        notify_subject_updated(subject)
        flash('Subject updated successfully!', 'success')
        return redirect(url_for('subjects'))
    
    return render_template('subjects/form.html', form=form, subject=subject, title='Edit Subject')

@app.route('/subjects/<int:id>/delete', methods=['POST'])
@login_required
def delete_subject(id):
    subject = Subject.query.get_or_404(id)
    subject.is_active = False
    subject.updated_at = datetime.utcnow()
    db.session.commit()
    
    flash('Subject deleted successfully!', 'success')
    return redirect(url_for('subjects'))

# Attendance Management Routes
@app.route('/attendance')
@login_required
@teacher_or_admin_required
def attendance():
    date_filter = request.args.get('date', datetime.now().date().isoformat())
    
    # Teachers can view all attendance records, editing restrictions handled elsewhere
    attendance_records = Attendance.query.filter_by(date=date_filter).all()
    students = Student.query.filter_by(status='active').all()
    
    return render_template('attendance/list.html', 
                         attendance_records=attendance_records,
                         students=students,
                         selected_date=date_filter)

@app.route('/attendance/record', methods=['GET', 'POST'])
@login_required
@teacher_or_admin_required
def record_attendance():
    form = AttendanceForm()
    
    # For teachers, restrict to students they teach
    if current_user.role == 'teacher':
        teacher_students = get_teacher_students(current_user.id)
        form.student_id.choices = [(s.id, s.full_name) for s in teacher_students]
    else:
        # Admin can record attendance for all students
        form.student_id.choices = [(s.id, s.full_name) for s in Student.query.filter_by(status='active').all()]
    
    if form.validate_on_submit():
        # Additional check for teachers
        if current_user.role == 'teacher':
            if not can_teacher_access_student(current_user.id, form.student_id.data):
                flash('You can only record attendance for students you teach.', 'error')
                return redirect(url_for('attendance'))
        
        # Determine period based on student's faculty
        student = Student.query.get(form.student_id.data)
        enrollment = Enrollment.query.filter_by(student_id=student.id).first()
        if enrollment and enrollment.subject:
            if enrollment.subject.code == 'BCA':
                period = 'morning'
            else:  # BSC-CSIT
                period = 'day'
        else:
            # Default fallback if no enrollment found
            period = 'morning'
        
        attendance = Attendance(
            student_id=form.student_id.data,
            date=form.date.data,
            status=form.status.data,
            period=period,
            notes=form.notes.data,
            recorded_by=current_user.id
        )
        db.session.add(attendance)
        db.session.commit()
        
        notify_attendance_recorded(attendance)
        flash('Attendance recorded successfully!', 'success')
        return redirect(url_for('attendance'))
    
    return render_template('attendance/form.html', form=form, title='Record Attendance')

@app.route('/attendance/<int:attendance_id>/edit', methods=['GET', 'POST'])
@login_required
@teacher_or_admin_required
def edit_attendance(attendance_id):
    attendance = Attendance.query.get_or_404(attendance_id)
    
    # Check if teacher can access this attendance record
    if current_user.role == 'teacher':
        if not can_teacher_access_student(current_user.id, attendance.student_id):
            flash('You can only edit attendance for students you teach.', 'error')
            return redirect(url_for('attendance'))
    
    form = AttendanceForm(obj=attendance)
    
    # For teachers, restrict to students they teach
    if current_user.role == 'teacher':
        teacher_students = get_teacher_students(current_user.id)
        form.student_id.choices = [(s.id, s.full_name) for s in teacher_students]
    else:
        # Admin can edit attendance for all students
        form.student_id.choices = [(s.id, s.full_name) for s in Student.query.filter_by(status='active').all()]
    
    if form.validate_on_submit():
        # Additional check for teachers
        if current_user.role == 'teacher':
            if not can_teacher_access_student(current_user.id, form.student_id.data):
                flash('You can only edit attendance for students you teach.', 'error')
                return redirect(url_for('attendance'))
        
        form.populate_obj(attendance)
        attendance.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        flash('Attendance updated successfully!', 'success')
        return redirect(url_for('attendance'))
    
    return render_template('attendance/form.html', form=form, attendance=attendance, title='Edit Attendance')

@app.route('/attendance/bulk', methods=['GET', 'POST'])
@login_required
@teacher_or_admin_required
def bulk_attendance():
    if request.method == 'POST':
        date = request.form.get('date')
        students = Student.query.filter_by(status='active').all()
        
        for student in students:
            status = request.form.get(f'status_{student.id}')
            if status:
                # Determine period based on student's faculty
                # Get student's enrolled faculty to determine period
                enrollment = Enrollment.query.filter_by(student_id=student.id).first()
                if enrollment and enrollment.subject:
                    if enrollment.subject.code == 'BCA':
                        period = 'morning'
                    else:  # BSC-CSIT
                        period = 'day'
                else:
                    # Default fallback if no enrollment found
                    period = 'morning'
                
                attendance = Attendance(
                    student_id=student.id,
                    date=datetime.strptime(date, '%Y-%m-%d').date(),
                    status=status,
                    period=period,
                    recorded_by=current_user.id
                )
                db.session.add(attendance)
        
        db.session.commit()
        flash('Bulk attendance recorded successfully!', 'success')
        return redirect(url_for('attendance'))
    
    students = Student.query.filter_by(status='active').order_by(Student.last_name, Student.first_name).all()
    today = datetime.now().strftime('%Y-%m-%d')
    return render_template('attendance/bulk_form.html', students=students, today=today)

# Grade Management Routes
@app.route('/grades')
@login_required
@teacher_or_admin_required
def grades():
    student_filter = request.args.get('student_id', type=int)
    subject_filter = request.args.get('subject_id', type=int)
    
    query = Grade.query.join(Student).join(Subject)
    
    # Teachers can view all grades, restriction is only on editing
    # (editing restrictions are handled at individual grade route level)
    
    if student_filter:
        query = query.filter(Grade.student_id == student_filter)
    if subject_filter:
        query = query.filter(Grade.subject_id == subject_filter)
    
    grades = query.order_by(desc(Grade.date_recorded)).all()
    
    # Process grades to add percentage and remarks
    processed_grades = []
    for grade in grades:
        # Calculate percentage (assuming grade_value is out of 100)
        percentage = grade.grade_value if grade.grade_value else 0
        
        # Determine remarks (pass/fail)
        remarks = "Pass" if percentage >= 40 else "Fail"
        
        # Use semester directly (now stored as 1st, 2nd, 3rd, etc.)
        semester = grade.semester
        
        processed_grades.append({
            'id': grade.id,
            'student_name': grade.student.full_name,
            'student_id': grade.student.student_id,
            'faculty': grade.subject.code,  # BSC-CSIT or BCA
            'percentage': percentage,
            'letter_grade': grade.letter_grade,
            'remarks': remarks,
            'semester': semester,
            'date_recorded': grade.date_recorded
        })
    
    students = Student.query.filter_by(status='active').all()
    subjects = Subject.query.filter_by(is_active=True).all()
    
    return render_template('grades/list.html', 
                         processed_grades=processed_grades,
                         students=students,
                         subjects=subjects,
                         student_filter=student_filter,
                         subject_filter=subject_filter)

@app.route('/grades/add', methods=['GET', 'POST'])
@login_required
@teacher_or_admin_required
def add_grade():
    form = GradeForm()
    
    # Set default academic year if not provided
    if not form.academic_year.data:
        form.academic_year.data = '2024-25'
    
    # For teachers, restrict to students they teach and subjects they teach
    if current_user.role == 'teacher':
        teacher_students = get_teacher_students(current_user.id)
        form.student_id.choices = [(s.id, s.full_name) for s in teacher_students]
        
        # Get subjects that this teacher teaches
        teacher = Teacher.query.filter_by(user_id=current_user.id).first()
        if teacher:
            teacher_subject_ids = db.session.query(SubjectTeacher.subject_id).filter_by(teacher_id=teacher.id).subquery()
            teacher_subjects = Subject.query.filter(Subject.id.in_(teacher_subject_ids), Subject.is_active == True).all()
            form.subject_id.choices = [(s.id, s.name) for s in teacher_subjects]
        else:
            form.subject_id.choices = []
    else:
        # Admin can access all students and subjects
        form.student_id.choices = [(s.id, s.full_name) for s in Student.query.filter_by(status='active').all()]
        form.subject_id.choices = [(s.id, s.name) for s in Subject.query.filter_by(is_active=True).all()]
    
    if form.validate_on_submit():
        # Additional check for teachers
        if current_user.role == 'teacher':
            if not can_teacher_access_student(current_user.id, form.student_id.data):
                flash('You can only add grades for students you teach.', 'error')
                return redirect(url_for('grades'))
        
        grade = Grade(
            student_id=form.student_id.data,
            subject_id=form.subject_id.data,
            grade_value=form.grade_value.data,
            grade_type=form.grade_type.data,
            semester=form.semester.data,
            academic_year=form.academic_year.data or '2024-25',
            date_recorded=form.date_recorded.data,
            comments=form.comments.data
        )
        
        # Calculate letter grade
        grade.letter_grade = form.letter_grade.data if form.letter_grade.data else grade.calculate_letter_grade()
        
        try:
            db.session.add(grade)
            db.session.commit()
            
            # Update student GPA
            student = Student.query.get(grade.student_id)
            if student:
                student.gpa = student.calculate_gpa()
                db.session.commit()
            
            flash('Grade added successfully!', 'success')
            return redirect(url_for('grades'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error saving grade: {str(e)}', 'error')
            return render_template('grades/form.html', form=form, title='Add Grade')
    
    return render_template('grades/form.html', form=form, title='Add Grade')

@app.route('/grades/<int:grade_id>/edit', methods=['GET', 'POST'])
@login_required
@teacher_or_admin_required
def edit_grade(grade_id):
    grade = Grade.query.get_or_404(grade_id)
    
    # Check if teacher can access this grade
    if current_user.role == 'teacher':
        if not can_teacher_access_student(current_user.id, grade.student_id):
            flash('You can only edit grades for students you teach.', 'error')
            return redirect(url_for('grades'))
    
    form = GradeForm(obj=grade)
    
    # For teachers, restrict to students they teach and subjects they teach
    if current_user.role == 'teacher':
        teacher_students = get_teacher_students(current_user.id)
        form.student_id.choices = [(s.id, s.full_name) for s in teacher_students]
        
        # Get subjects that this teacher teaches
        teacher = Teacher.query.filter_by(user_id=current_user.id).first()
        if teacher:
            teacher_subject_ids = db.session.query(SubjectTeacher.subject_id).filter_by(teacher_id=teacher.id).subquery()
            teacher_subjects = Subject.query.filter(Subject.id.in_(teacher_subject_ids), Subject.is_active == True).all()
            form.subject_id.choices = [(s.id, s.name) for s in teacher_subjects]
        else:
            form.subject_id.choices = []
    else:
        # Admin can access all students and subjects
        form.student_id.choices = [(s.id, s.full_name) for s in Student.query.filter_by(status='active').all()]
        form.subject_id.choices = [(s.id, s.name) for s in Subject.query.filter_by(is_active=True).all()]
    
    if form.validate_on_submit():
        # Additional check for teachers
        if current_user.role == 'teacher':
            if not can_teacher_access_student(current_user.id, form.student_id.data):
                flash('You can only edit grades for students you teach.', 'error')
                return redirect(url_for('grades'))
        
        form.populate_obj(grade)
        grade.letter_grade = grade.calculate_letter_grade()
        grade.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        # Update student GPA
        student = Student.query.get(grade.student_id)
        if student:
            student.gpa = student.calculate_gpa()
            db.session.commit()
        
        flash('Grade updated successfully!', 'success')
        return redirect(url_for('grades'))
    
    return render_template('grades/form.html', form=form, grade=grade, title='Edit Grade')

# Reports Route
@app.route('/reports')
@login_required
@teacher_or_admin_required
def reports():
    return render_template('reports/index.html')

# Settings Route
@app.route('/settings')
@login_required
@admin_required
def settings():
    return render_template('settings/index.html')

# Error Handlers
@app.errorhandler(403)
def forbidden_error(error):
    flash('Access denied. You do not have permission to access this resource.', 'error')
    return redirect(url_for('dashboard'))

@app.errorhandler(404)
def not_found_error(error):
    return render_template('errors/404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return render_template('errors/500.html'), 500

# Student Portal Routes
@app.route('/student/dashboard')
@login_required
def student_dashboard():
    # Check if logged in as student
    if not hasattr(current_user, 'student_id'):
        flash('Access denied. Student portal only.', 'danger')
        return redirect(url_for('dashboard'))
    
    student = current_user
    
    # Get student's grades
    recent_grades = student.grades.order_by(desc(Grade.date_recorded)).limit(10).all()
    
    # Get student's attendance
    attendance_rate = student.get_attendance_rate(30)
    recent_attendance = student.attendance_records.order_by(desc(Attendance.date)).limit(10).all()
    
    # Get student's enrollments
    enrollments = student.enrollments.filter_by(status='enrolled').all()
    
    # Get student's assignments (ordered by submission date)
    recent_assignments = student.assignments.order_by(desc(Assignment.created_at)).limit(5).all()
    
    return render_template('student/dashboard.html',
                         student=student,
                         recent_grades=recent_grades,
                         attendance_rate=attendance_rate,
                         recent_attendance=recent_attendance,
                         enrollments=enrollments,
                         recent_assignments=recent_assignments)

@app.route('/student/profile')
@login_required
def student_profile():
    # Check if logged in as student
    if current_user.role != 'student':
        flash('Access denied. Student portal only.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Show search form instead of direct profile
    return render_template('student/profile.html')

@app.route('/student/search-profile', methods=['POST'])
@login_required
@student_required
def student_search_profile():
    """Search for student profile by name and ID"""
    search_term = request.form.get('search_term', '').strip()
    
    if not search_term:
        flash('Please enter your name or student ID', 'error')
        return redirect(url_for('student_profile'))
    
    # Search by ID or name
    student = None
    if search_term.startswith('STU') or search_term.isdigit():
        student = Student.query.filter_by(student_id=search_term, status='active').first()
    else:
        # Search by name - need to check first_name and last_name separately since full_name is a property
        student = Student.query.filter(
            db.or_(
                Student.first_name.ilike(f'%{search_term}%'),
                Student.last_name.ilike(f'%{search_term}%'),
                db.func.concat(Student.first_name, ' ', Student.last_name).ilike(f'%{search_term}%')
            ),
            Student.status == 'active'
        ).first()
    
    if not student:
        flash('Student not found. Please check your name or ID.', 'error')
        return redirect(url_for('student_profile'))
        
    return render_template('student/profile.html', student=student)

@app.route('/student/profile/<student_id>/edit', methods=['GET', 'POST'])
@login_required
@student_required
def student_edit_profile_with_id(student_id):
    """Allow students to edit their own profile (limited fields only) with student ID"""
    # Check if logged in as student
    if current_user.role != 'student':
        flash('Access denied. Student portal only.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Get the Student record by student_id
    student = Student.query.filter_by(student_id=student_id, status='active').first()
    if not student:
        flash('Student record not found.', 'error')
        return redirect(url_for('student_profile'))
    
    if request.method == 'POST':
        try:
            # Students can only edit these specific fields
            student.phone = request.form.get('phone', student.phone)
            student.address = request.form.get('address', student.address)
            student.emergency_contact = request.form.get('emergency_contact', student.emergency_contact)
            student.emergency_phone = request.form.get('emergency_phone', student.emergency_phone)
            
            # Handle profile photo upload if provided
            if 'profile_photo' in request.files:
                file = request.files['profile_photo']
                if file and file.filename and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    # Create unique filename
                    filename = f"student_{student.id}_{int(time.time())}_{filename}"
                    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                    file.save(file_path)
                    student.photo_filename = filename
            
            db.session.commit()
            flash('Profile updated successfully!', 'success')
            return redirect(url_for('student_search_profile'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating profile: {str(e)}', 'error')
    
    return render_template('student/edit_profile.html', student=student)

@app.route('/student/grades')
@login_required
def student_grades():
    # Check if logged in as student
    if current_user.role != 'student':
        flash('Access denied. Student portal only.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Get the actual Student record linked by email
    student = Student.query.filter_by(email=current_user.email).first()
    if not student:
        flash('Student record not found.', 'error')
        return redirect(url_for('dashboard'))
    
    # Get all grades grouped by subject
    grades_by_subject = {}
    for grade in student.grades.order_by(desc(Grade.date_recorded)).all():
        subject_name = grade.subject.name if grade.subject else "Unknown"
        if subject_name not in grades_by_subject:
            grades_by_subject[subject_name] = []
        grades_by_subject[subject_name].append(grade)
    
    return render_template('student/grades.html',
                         student=student,
                         grades_by_subject=grades_by_subject)

@app.route('/student/attendance')
@login_required
def student_attendance():
    # Check if logged in as student
    if current_user.role != 'student':
        flash('Access denied. Student portal only.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Get the actual Student record linked by email
    student = Student.query.filter_by(email=current_user.email).first()
    if not student:
        flash('Student record not found.', 'error')
        return redirect(url_for('dashboard'))
    
    # Get attendance records with pagination
    page = request.args.get('page', 1, type=int)
    attendance_pagination = student.attendance_records.order_by(desc(Attendance.date)).paginate(
        page=page, per_page=20, error_out=False
    )
    
    # Calculate stats
    total_records = student.attendance_records.count()
    present_count = student.attendance_records.filter_by(status='present').count()
    absent_count = student.attendance_records.filter_by(status='absent').count()
    late_count = student.attendance_records.filter_by(status='late').count()
    
    return render_template('student/attendance.html',
                         student=student,
                         attendance=attendance_pagination.items,
                         pagination=attendance_pagination,
                         total_records=total_records,
                         present_count=present_count,
                         absent_count=absent_count,
                         late_count=late_count)

@app.route('/student/change-password', methods=['GET', 'POST'])
@login_required  
def student_change_password():
    """Allow students to change their own password"""
    # Check if logged in as student
    if current_user.role != 'student':
        flash('Access denied. Student portal only.', 'danger')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        
        # Verify current password
        if not current_user.check_password(current_password):
            flash('Current password is incorrect.', 'error')
            return render_template('student/change_password.html')
        
        # Check if new passwords match
        if new_password != confirm_password:
            flash('New passwords do not match.', 'error')
            return render_template('student/change_password.html')
        
        # Check password strength
        if len(new_password) < 6:
            flash('Password must be at least 6 characters long.', 'error')
            return render_template('student/change_password.html')
        
        try:
            current_user.set_password(new_password)
            db.session.commit()
            flash('Password changed successfully!', 'success')
            return redirect(url_for('student_profile'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error changing password: {str(e)}', 'error')
    
    return render_template('student/change_password.html')

# Initialize Database
def init_db():
    """Initialize database with error handling for serverless deployment"""
    try:
        with app.app_context():
            # Only create tables, don't modify data in serverless
            db.create_all()
            
            # Skip data initialization in serverless environment
            if not os.getenv('VERCEL'):
                # Create default admin user if not exists
                if not User.query.filter_by(username='admin').first():
                    admin = User(username='admin', email='admin@school.com', role='admin')
                    admin.set_password('admin123')
                    db.session.add(admin)
                
                # Create default teacher users if not exist
                if not User.query.filter_by(username='teacher').first():
                    teacher = User(username='teacher', email='teacher@school.com', role='teacher')
                    teacher.set_password('teacher123')
                    db.session.add(teacher)
                
                # Create student user if not exists
                if not User.query.filter_by(username='student').first():
                    student_user = User(username='student', email='student@school.com', role='student')
                    student_user.set_password('student123')
                    db.session.add(student_user)
                    
                # Create sample subjects if not exist
                if not Subject.query.first():
                    subjects = [
                        Subject(name='Mathematics', code='MATH101', credits=4, department='Mathematics'),
                        Subject(name='English', code='ENG101', credits=3, department='English'),
                        Subject(name='Science', code='SCI101', credits=4, department='Science'),
                        Subject(name='History', code='HIST101', credits=3, department='Social Studies'),
                        Subject(name='Physical Education', code='PE101', credits=2, department='PE')
                    ]
                    for subject in subjects:
                        db.session.add(subject)
                
                # Create current academic year if not exists
                if not AcademicYear.query.filter_by(is_current=True).first():
                    current_year = AcademicYear(
                        year='2024-25',
                        start_date=datetime(2024, 8, 1).date(),
                        end_date=datetime(2025, 7, 31).date(),
                        is_current=True
                    )
                    db.session.add(current_year)
                
                db.session.commit()
                print("Database initialized successfully!")
    except Exception as e:
        print(f"Database initialization warning: {e}")
        # Don't fail startup due to database issues
        print("Login credentials:")
        print("  Admin: admin / admin123")
        print("  Teacher: teacher / teacher123")
        print("  Student: student / student123")
# Student-specific routes for viewing their own data
@app.route('/student/my-grades')
@login_required
@student_required
def student_my_grades():
    # For now, allow students to search by name/ID
    # Later can be enhanced to show grades for logged-in student
    return render_template('student/grades.html')

@app.route('/student/search-grades', methods=['POST'])
@login_required
@student_required
def student_search_grades():
    search_term = request.form.get('search_term', '').strip()
    
    if not search_term:
        flash('Please enter your name or student ID', 'error')
        return redirect(url_for('student_my_grades'))
    
    # Search by ID or name
    student = None
    if search_term.startswith('STU') or search_term.isdigit():
        student = Student.query.filter_by(student_id=search_term, status='active').first()
    else:
        # Search by name - need to check first_name and last_name separately since full_name is a property
        student = Student.query.filter(
            db.or_(
                Student.first_name.ilike(f'%{search_term}%'),
                Student.last_name.ilike(f'%{search_term}%'),
                db.func.concat(Student.first_name, ' ', Student.last_name).ilike(f'%{search_term}%')
            ),
            Student.status == 'active'
        ).first()
    
    if not student:
        flash('Student not found. Please check your name or ID.', 'error')
        return redirect(url_for('student_my_grades'))
    
    grades = Grade.query.filter_by(student_id=student.id).join(Subject).all()
    return render_template('student/grades.html', student=student, grades=grades)

@app.route('/student/my-attendance')
@login_required
@student_required
def student_my_attendance():
    return render_template('student/attendance.html')

@app.route('/student/search-attendance', methods=['POST'])
@login_required
@student_required
def student_search_attendance():
    search_term = request.form.get('search_term', '').strip()
    
    if not search_term:
        flash('Please enter your name or student ID', 'error')
        return redirect(url_for('student_my_attendance'))
    
    # Search by ID or name
    student = None
    if search_term.startswith('STU') or search_term.isdigit():
        student = Student.query.filter_by(student_id=search_term, status='active').first()
    else:
        # Search by name - need to check first_name and last_name separately since full_name is a property
        student = Student.query.filter(
            db.or_(
                Student.first_name.ilike(f'%{search_term}%'),
                Student.last_name.ilike(f'%{search_term}%'),
                db.func.concat(Student.first_name, ' ', Student.last_name).ilike(f'%{search_term}%')
            ),
            Student.status == 'active'
        ).first()
    
    if not student:
        flash('Student not found. Please check your name or ID.', 'error')
        return redirect(url_for('student_my_attendance'))
    
    # Get all attendance for percentage calculation
    all_attendance = Attendance.query.filter_by(student_id=student.id).all()
    total_days = len(all_attendance)
    present_days = len([a for a in all_attendance if a.status == 'present'])
    attendance_percentage = (present_days / total_days * 100) if total_days > 0 else 0
    
    # Get only last 10 days for display
    attendance = Attendance.query.filter_by(student_id=student.id).order_by(Attendance.date.desc()).limit(10).all()
    
    return render_template('student/attendance.html', 
                         student=student, 
                         attendance=attendance,
                         total_days=total_days,
                         present_days=present_days,
                         attendance_percentage=attendance_percentage)

@app.route('/student/my-fees')
@login_required
@student_required
def student_my_fees():
    """Student fee status lookup page"""
    return render_template('student/fees.html')

@app.route('/student/search-fees', methods=['POST'])
@login_required
@student_required
def student_search_fees():
    """Search for student fee status by name and ID"""
    search_term = request.form.get('search_term', '').strip()
    
    if not search_term:
        flash('Please enter your name or student ID', 'error')
        return redirect(url_for('student_my_fees'))
    
    # Search by ID or name
    student = None
    if search_term.startswith('STU') or search_term.isdigit():
        student = Student.query.filter_by(student_id=search_term, status='active').first()
    else:
        # Search by name - need to check first_name and last_name separately since full_name is a property
        student = Student.query.filter(
            db.or_(
                Student.first_name.ilike(f'%{search_term}%'),
                Student.last_name.ilike(f'%{search_term}%'),
                db.func.concat(Student.first_name, ' ', Student.last_name).ilike(f'%{search_term}%')
            ),
            Student.status == 'active'
        ).first()
    
    if not student:
        flash('Student not found. Please check your name or ID.', 'error')
        return redirect(url_for('student_my_fees'))
    
    # Get student's faculty
    enrollment = Enrollment.query.filter_by(student_id=student.id, status='enrolled').first()
    if not enrollment:
        flash('No active enrollment found.', 'error')
        return redirect(url_for('student_my_fees'))
    
    # Get applicable fee structures for this student's faculty
    fee_structures = FeeStructure.query.filter_by(
        faculty_id=enrollment.subject_id,
        is_active=True
    ).all()
    
    # Calculate fee summary for each structure
    fee_summary = []
    total_fees = 0
    total_paid = 0
    total_pending = 0
    
    for structure in fee_structures:
        # Get payments for this fee structure
        payments = FeePayment.query.filter_by(
            student_id=student.id,
            fee_structure_id=structure.id
        ).all()
        
        paid_amount = sum(payment.amount for payment in payments)
        pending_amount = structure.amount - paid_amount
        
        fee_summary.append({
            'structure': structure,
            'paid_amount': paid_amount,
            'pending_amount': pending_amount,
            'payments': payments
        })
        
        total_fees += structure.amount
        total_paid += paid_amount
        total_pending += pending_amount
    
    # Overall status
    if total_pending == 0:
        fee_status = 'Paid'
        status_color = 'success'
    elif total_paid > 0:
        fee_status = 'Partially Paid'
        status_color = 'warning'
    else:
        fee_status = 'Pending'
        status_color = 'danger'
    
    return render_template('student/fees.html', 
                         student=student, 
                         fee_summary=fee_summary,
                         total_fees=total_fees,
                         total_paid=total_paid,
                         total_pending=total_pending,
                         fee_status=fee_status,
                         status_color=status_color,
                         faculty=enrollment.subject)

# Fee Management Routes
@app.route('/fees')
@login_required
def fees():
    """Fee management main page"""
    if current_user.role == 'student':
        return redirect(url_for('student_fees'))
    elif current_user.role == 'teacher':
        return redirect(url_for('teacher_fees'))
    else:  # admin
        return redirect(url_for('admin_fees'))

@app.route('/fees/student')
@login_required  
def student_fees():
    """Student fee view - can only view their own fees"""
    if current_user.role != 'student':
        flash('Access denied. Students only.', 'error')
        return redirect(url_for('dashboard'))
    
    # Get current student
    student = Student.query.filter_by(email=current_user.email).first()
    if not student:
        flash('Student record not found.', 'error')
        return redirect(url_for('dashboard'))
    
    # Get student's faculty
    enrollment = Enrollment.query.filter_by(student_id=student.id, status='enrolled').first()
    if not enrollment:
        flash('No active enrollment found.', 'error')
        return redirect(url_for('dashboard'))
    
    # Get applicable fee structures for this student
    fee_structures = FeeStructure.query.filter_by(
        faculty_id=enrollment.subject_id,
        is_active=True
    ).all()
    
    # Calculate fee summary for each structure
    fee_summary = []
    total_fees = 0
    total_paid = 0
    total_pending = 0
    
    for fee_structure in fee_structures:
        payments = FeePayment.query.filter_by(
            student_id=student.id,
            fee_structure_id=fee_structure.id
        ).all()
        
        paid_amount = sum(payment.amount_paid for payment in payments)
        pending_amount = fee_structure.amount - paid_amount
        
        fee_summary.append({
            'fee_structure': fee_structure,
            'total_amount': fee_structure.amount,
            'paid_amount': paid_amount,
            'pending_amount': max(0, pending_amount),
            'payments': payments,
            'is_overdue': fee_structure.due_date and fee_structure.due_date < datetime.now().date() and pending_amount > 0
        })
        
        total_fees += fee_structure.amount
        total_paid += paid_amount
        total_pending += max(0, pending_amount)
    
    return render_template('fees/student.html',
                         student=student,
                         fee_summary=fee_summary,
                         total_fees=total_fees,
                         total_paid=total_paid,
                         total_pending=total_pending)

@app.route('/fees/teacher')
@login_required
def teacher_fees():
    """Teacher fee view - can view all students fee status"""
    if current_user.role != 'teacher':
        flash('Access denied. Teachers only.', 'error')
        return redirect(url_for('dashboard'))
    
    # Get all fee structures
    fee_structures = FeeStructure.query.filter_by(is_active=True).order_by(FeeStructure.faculty_id, FeeStructure.semester).all()
    
    # Get all active students with their fee calculations
    students = Student.query.filter_by(status='active').all()
    student_fees = []
    
    for student in students:
        # Get student's enrollment to find faculty
        enrollment = Enrollment.query.filter_by(student_id=student.id, status='enrolled').first()
        if not enrollment:
            continue
            
        # Get applicable fee structures for this student
        student_fee_structures = FeeStructure.query.filter_by(
            faculty_id=enrollment.subject_id,
            is_active=True
        ).all()
        
        total_fees = sum(fs.amount for fs in student_fee_structures)
        total_paid = 0
        has_overdue = False
        
        for fee_structure in student_fee_structures:
            payments = FeePayment.query.filter_by(
                student_id=student.id,
                fee_structure_id=fee_structure.id
            ).all()
            total_paid += sum(p.amount_paid for p in payments)
            
            # Check if overdue
            if fee_structure.due_date and fee_structure.due_date < datetime.now().date():
                paid_for_structure = sum(p.amount_paid for p in payments)
                if paid_for_structure < fee_structure.amount:
                    has_overdue = True
        
        total_pending = max(0, total_fees - total_paid)
        
        student_fees.append({
            'student': student,
            'total_fees': total_fees,
            'total_paid': total_paid,
            'total_pending': total_pending,
            'has_overdue': has_overdue
        })
    
    return render_template('fees/teacher.html',
                         fee_structures=fee_structures,
                         student_fees=student_fees)

@app.route('/fees/admin')
@login_required
def admin_fees():
    """Admin fee management - full access to fee structures and payments"""
    if current_user.role != 'admin':
        flash('Access denied. Admins only.', 'error')
        return redirect(url_for('dashboard'))
    
    # Get all fee structures
    fee_structures = FeeStructure.query.order_by(FeeStructure.faculty_id, FeeStructure.semester).all()
    
    # Calculate proper summary statistics
    # Get all students and their enrollments
    students = Student.query.filter_by(status='active').all()
    total_fees_defined = 0
    total_payments = db.session.query(func.sum(FeePayment.amount_paid)).scalar() or 0
    
    # Calculate total fees based on student enrollments
    for student in students:
        enrollment = Enrollment.query.filter_by(student_id=student.id, status='enrolled').first()
        if enrollment:
            student_fee_structures = FeeStructure.query.filter_by(
                faculty_id=enrollment.subject_id,
                is_active=True
            ).all()
            total_fees_defined += sum(fs.amount for fs in student_fee_structures)
    
    pending_payments = max(0, total_fees_defined - total_payments)
    
    # Get recent payments with relationships loaded
    recent_payments = FeePayment.query.join(Student).join(FeeStructure).order_by(desc(FeePayment.created_at)).limit(10).all()
    
    return render_template('fees/admin.html',
                         fee_structures=fee_structures,
                         total_fees_defined=total_fees_defined,
                         total_payments=total_payments,
                         pending_payments=pending_payments,
                         recent_payments=recent_payments)

@app.route('/fees/structure/add', methods=['GET', 'POST'])
@login_required
def add_fee_structure():
    """Admin only - Add new fee structure"""
    if current_user.role != 'admin':
        flash('Access denied. Admins only.', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        try:
            fee_structure = FeeStructure(
                name=request.form['name'],
                description=request.form.get('description', ''),
                amount=float(request.form['amount']),
                fee_type=request.form['fee_type'],
                faculty_id=int(request.form['faculty_id']) if request.form['faculty_id'] else None,
                semester=request.form.get('semester') if request.form.get('semester') else None,
                academic_year=request.form['academic_year'],
                due_date=datetime.strptime(request.form['due_date'], '%Y-%m-%d').date() if request.form.get('due_date') else None
            )
            
            db.session.add(fee_structure)
            db.session.commit()
            flash('Fee structure added successfully!', 'success')
            return redirect(url_for('admin_fees'))
        except Exception as e:
            flash(f'Error adding fee structure: {str(e)}', 'error')
    
    # Get faculties for dropdown
    faculties = Subject.query.all()
    return render_template('fees/add_structure.html', faculties=faculties)

@app.route('/fees/payment/add', methods=['GET', 'POST'])
@login_required
def add_fee_payment():
    """Admin only - Record new payment"""
    if current_user.role != 'admin':
        flash('Access denied. Admins only.', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        try:
            # Generate unique receipt number
            import uuid
            receipt_number = f"REC{datetime.now().strftime('%Y%m%d')}{str(uuid.uuid4())[:8].upper()}"
            
            payment = FeePayment(
                student_id=int(request.form['student_id']),
                fee_structure_id=int(request.form['fee_structure_id']),
                amount_paid=float(request.form['amount_paid']),
                payment_date=datetime.strptime(request.form['payment_date'], '%Y-%m-%d').date(),
                payment_method=request.form['payment_method'],
                transaction_id=request.form.get('transaction_id', ''),
                receipt_number=receipt_number,
                remarks=request.form.get('remarks', ''),
                recorded_by=current_user.id
            )
            
            db.session.add(payment)
            db.session.commit()
            flash('Payment recorded successfully!', 'success')
            return redirect(url_for('admin_fees'))
        except Exception as e:
            flash(f'Error recording payment: {str(e)}', 'error')
    
    # Get data for dropdowns
    students = Student.query.filter_by(status='active').all()
    fee_structures = FeeStructure.query.filter_by(is_active=True).all()
    today = datetime.now().date()
    return render_template('fees/add_payment.html', 
                         students=students, 
                         fee_structures=fee_structures,
                         today=today)

@app.route('/fees/structure/toggle/<int:structure_id>')
@login_required
def toggle_fee_structure(structure_id):
    """Admin only - Toggle fee structure active status"""
    if current_user.role != 'admin':
        flash('Access denied. Admins only.', 'error')
        return redirect(url_for('dashboard'))
    
    fee_structure = FeeStructure.query.get_or_404(structure_id)
    fee_structure.is_active = not fee_structure.is_active
    fee_structure.updated_at = datetime.utcnow()
    
    try:
        db.session.commit()
        status = 'activated' if fee_structure.is_active else 'deactivated'
        flash(f'Fee structure {status} successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating fee structure: {str(e)}', 'error')
    
    return redirect(url_for('admin_fees'))

@app.route('/fees/structure/edit/<int:structure_id>', methods=['GET', 'POST'])
@login_required
def edit_fee_structure(structure_id):
    """Admin only - Edit fee structure"""
    if current_user.role != 'admin':
        flash('Access denied. Admins only.', 'error')
        return redirect(url_for('dashboard'))
    
    fee_structure = FeeStructure.query.get_or_404(structure_id)
    
    if request.method == 'POST':
        try:
            fee_structure.name = request.form['name']
            fee_structure.description = request.form.get('description', '')
            fee_structure.amount = float(request.form['amount'])
            fee_structure.fee_type = request.form['fee_type']
            fee_structure.faculty_id = int(request.form['faculty_id']) if request.form['faculty_id'] else None
            fee_structure.semester = request.form.get('semester') if request.form.get('semester') else None
            fee_structure.academic_year = request.form['academic_year']
            fee_structure.due_date = datetime.strptime(request.form['due_date'], '%Y-%m-%d').date() if request.form.get('due_date') else None
            fee_structure.updated_at = datetime.utcnow()
            
            db.session.commit()
            flash('Fee structure updated successfully!', 'success')
            return redirect(url_for('admin_fees'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating fee structure: {str(e)}', 'error')
    
    # Get faculties for dropdown
    faculties = Subject.query.all()
    return render_template('fees/edit_structure.html', fee_structure=fee_structure, faculties=faculties)

@app.route('/fees/reports')
@login_required
def fee_reports():
    """Fee reports for admins and teachers"""
    if current_user.role not in ['admin', 'teacher']:
        flash('Access denied.', 'error')
        return redirect(url_for('dashboard'))
    
    return render_template('fees/reports.html')

@app.route('/api/fees/export/<export_type>')
@login_required
def export_fee_data(export_type):
    """Export fee data as CSV or JSON"""
    if current_user.role not in ['admin', 'teacher']:
        flash('Access denied.', 'error')
        return redirect(url_for('dashboard'))
    
    # Get fee data based on user role
    if current_user.role == 'teacher':
        teacher = Teacher.query.filter_by(email=current_user.email).first()
        if teacher:
            assigned_students = get_teacher_students(teacher.id)
            student_ids = [s.id for s in assigned_students]
            fee_data = db.session.query(
                FeePayment, Student, FeeStructure
            ).join(Student).join(FeeStructure).filter(
                Student.id.in_(student_ids)
            ).all()
        else:
            fee_data = []
    else:  # admin
        fee_data = db.session.query(
            FeePayment, Student, FeeStructure
        ).join(Student).join(FeeStructure).all()
    
    # Prepare export data
    export_data = []
    for payment, student, fee_structure in fee_data:
        export_data.append({
            'student_id': student.student_id,
            'student_name': f"{student.first_name} {student.last_name}",
            'faculty': fee_structure.faculty.name if fee_structure.faculty else 'N/A',
            'fee_name': fee_structure.name,
            'fee_amount': fee_structure.amount,
            'amount_paid': payment.amount_paid,
            'payment_date': payment.payment_date.isoformat(),
            'payment_method': payment.payment_method,
            'receipt_number': payment.receipt_number,
            'semester': fee_structure.semester,
            'academic_year': fee_structure.academic_year
        })
    
    if export_type == 'csv':
        import csv
        import io
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=export_data[0].keys() if export_data else [])
        writer.writeheader()
        writer.writerows(export_data)
        
        response = Response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = 'attachment; filename=fee_payments.csv'
        return response
    
    elif export_type == 'json':
        response = jsonify(export_data)
        response.headers['Content-Disposition'] = 'attachment; filename=fee_payments.json'
        return response
    
    flash('Invalid export format.', 'error')
    return redirect(url_for('fee_reports'))

if __name__ == '__main__':
    init_db()
    socketio.run(app, debug=True, host='0.0.0.0', port=5000)
else:
    # Vercel serverless handler
    init_db()

# Export app for Vercel
application = app
